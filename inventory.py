"""
inventory.xlsx への商品追記ヘルパー。
これまで手作業で繰り返してきた openpyxl 追記処理をモジュール化。
A=管理番号 B=商品名 C=カテゴリ D=コンディション E=仕入れ値円 F=販売価格円(USD×160.08)
G==F-E H==IFERROR(G/F) I=在庫 J=出品 K=売約 L==MAX(I-J-K,0) M=eBay ID N=URL O=ステータス Q=備考
"""

from pathlib import Path
from copy import copy
from openpyxl import load_workbook

XLSX = Path(__file__).resolve().parent.parent / "inventory.xlsx"
RATE = 160.08  # ¥/$ 換算レート


def next_item_no(ws) -> str:
    """既存の最大 ITEM-XXX の次番号を返す。"""
    maxn = 0
    for row in range(2, ws.max_row + 1):
        v = ws.cell(row=row, column=1).value
        if isinstance(v, str) and v.startswith("ITEM-"):
            try:
                maxn = max(maxn, int(v.split("-")[1]))
            except (IndexError, ValueError):
                pass
    return f"ITEM-{maxn + 1:03d}"


def append_item(name, category, condition, price_usd, ebay_id,
                cost_yen=None, note="", xlsx_path=XLSX):
    """新規商品を1行追加し、付与した ITEM 番号を返す。書式は直前行からコピー。"""
    wb = load_workbook(xlsx_path)
    ws = wb["在庫リスト"]
    r = ws.max_row + 1
    prev = r - 1
    item_no = next_item_no(ws)

    # 直前行から書式コピー
    for c in range(1, 18):
        cell = ws.cell(row=r, column=c)
        src = ws.cell(row=prev, column=c)
        cell.font = copy(src.font)
        cell.border = copy(src.border)
        cell.fill = copy(src.fill)
        cell.number_format = src.number_format
        cell.alignment = copy(src.alignment)

    ws.cell(row=r, column=1).value = item_no
    ws.cell(row=r, column=2).value = name
    ws.cell(row=r, column=3).value = category
    ws.cell(row=r, column=4).value = condition
    if cost_yen is not None:
        ws.cell(row=r, column=5).value = cost_yen
    ws.cell(row=r, column=6).value = round(price_usd * RATE)
    ws.cell(row=r, column=7).value = f"=F{r}-E{r}"
    ws.cell(row=r, column=8).value = f'=IFERROR(G{r}/F{r},"")'
    ws.cell(row=r, column=9).value = 1
    ws.cell(row=r, column=10).value = 1
    ws.cell(row=r, column=11).value = 0
    ws.cell(row=r, column=12).value = f"=MAX(I{r}-J{r}-K{r},0)"
    ws.cell(row=r, column=13).value = ebay_id
    ws.cell(row=r, column=14).value = f"https://www.ebay.com/itm/{ebay_id}"
    ws.cell(row=r, column=15).value = "出品中"
    ws.cell(row=r, column=17).value = note

    wb.save(xlsx_path)
    return item_no
